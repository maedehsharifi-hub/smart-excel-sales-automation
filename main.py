import pandas as pd
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_sales_data(input_folder):
    """Read, validate, and combine all Excel files from the input folder."""

    input_folder = Path(input_folder)

    excel_files = [
        file
        for file in input_folder.glob("*.xlsx")
        if not file.name.startswith("~$")
    ]

    print("Excel files found:", excel_files)
    print("Number of files:", len(excel_files))

    if not excel_files:
        raise ValueError("No Excel files found in the input folder.")

    required_columns = [
        "Name",
        "Email",
        "Product",
        "Quantity",
        "Price",
    ]

    numeric_columns = [
        "Quantity",
        "Price",
    ]

    all_data = []

    for file in excel_files:
        data = pd.read_excel(file)

        missing_columns = [
            column
            for column in required_columns
            if column not in data.columns
        ]

        if missing_columns:
            raise ValueError(
                f"{file.name} is missing required columns: "
                f"{', '.join(missing_columns)}"
            )

        for column in numeric_columns:
            converted = pd.to_numeric(
                data[column],
                errors="coerce",
            )

            if converted.isna().sum() > data[column].isna().sum():
                raise ValueError(
                    f"{file.name} contains invalid numeric data "
                    f"in column: {column}"
                )

            data[column] = converted

        all_data.append(data)

    combined_data = pd.concat(
        all_data,
        ignore_index=True,
    )

    return combined_data


def clean_sales_data(data):
    """Clean sales data and calculate total for each order."""

    original_rows = len(data)

    duplicates_removed = data.duplicated().sum()

    cleaned_data = data.drop_duplicates().copy()

    missing_prices_removed = cleaned_data["Price"].isna().sum()

    cleaned_data = cleaned_data.dropna(
        subset=["Price"]
    ).copy()

    cleaned_data["Total"] = (
        cleaned_data["Quantity"]
        * cleaned_data["Price"]
    )

    return (
        cleaned_data,
        original_rows,
        duplicates_removed,
        missing_prices_removed,
    )


def create_summary(
    cleaned_data,
    original_rows,
    duplicates_removed,
    missing_prices_removed,
):
    """Create summary information for the sales report."""

    total_orders = len(cleaned_data)

    total_sales = cleaned_data["Total"].sum()

    average_order = cleaned_data["Total"].mean()

    highest_order = cleaned_data["Total"].max()

    summary = pd.DataFrame(
        {
            "Metric": [
                "Original rows",
                "Duplicates removed",
                "Missing prices removed",
                "Total orders",
                "Total sales",
                "Average order",
                "Highest order",
            ],
            "Value": [
                original_rows,
                duplicates_removed,
                missing_prices_removed,
                total_orders,
                total_sales,
                average_order,
                highest_order,
            ],
        }
    )

    return summary


def export_report(
    cleaned_data,
    summary,
    output_file,
):
    """Export cleaned data and summary to Excel."""

    output_path = Path(output_file)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(output_path) as writer:

        cleaned_data.to_excel(
            writer,
            sheet_name="Cleaned Data",
            index=False,
        )

        summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )


def format_excel_report(output_file):
    """Apply professional formatting to the Excel report."""

    workbook = load_workbook(output_file)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    summary_label_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    number_format = "#,##0.00"

    # -------------------------
    # General formatting
    # -------------------------

    for worksheet in workbook.worksheets:

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center"
            )

        for column_cells in worksheet.columns:

            longest_value = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            worksheet.column_dimensions[
                column_letter
            ].width = longest_value + 3

    # -------------------------
    # Cleaned Data
    # -------------------------

    cleaned_sheet = workbook["Cleaned Data"]

    cleaned_sheet.freeze_panes = "A2"

    cleaned_sheet.auto_filter.ref = (
        cleaned_sheet.dimensions
    )

    cleaned_headers = {
        cell.value: cell.column
        for cell in cleaned_sheet[1]
    }

    for column_name in ["Price", "Total"]:

        column_number = cleaned_headers.get(
            column_name
        )

        if column_number is not None:

            for row_number in range(
                2,
                cleaned_sheet.max_row + 1,
            ):

                cleaned_sheet.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = number_format

    # -------------------------
    # Summary
    # -------------------------

    summary_sheet = workbook["Summary"]

    for row_number in range(
        2,
        summary_sheet.max_row + 1,
    ):

        metric_cell = summary_sheet.cell(
            row=row_number,
            column=1,
        )

        value_cell = summary_sheet.cell(
            row=row_number,
            column=2,
        )

        metric_cell.font = Font(
            bold=True
        )

        metric_cell.fill = summary_label_fill

        if metric_cell.value in [
            "Total sales",
            "Average order",
            "Highest order",
        ]:
            value_cell.number_format = number_format

    # -------------------------
    # Final fixed column widths
    # -------------------------

    cleaned_sheet.column_dimensions["A"].width = 14
    cleaned_sheet.column_dimensions["B"].width = 25
    cleaned_sheet.column_dimensions["C"].width = 16
    cleaned_sheet.column_dimensions["D"].width = 14
    cleaned_sheet.column_dimensions["E"].width = 20
    cleaned_sheet.column_dimensions["F"].width = 20

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 18

    workbook.save(output_file)

    print("Excel formatting applied.")


def print_summary(summary):
    """Print the sales summary in the terminal."""

    print("\n----- SALES REPORT -----")

    for _, row in summary.iterrows():
        print(
            f"{row['Metric']}: "
            f"{row['Value']}"
        )


def main():

    input_folder = "input"

    output_file = (
        "output/final_sales_report.xlsx"
    )

    try:

        sales_data = load_sales_data(
            input_folder
        )

        (
            cleaned_data,
            original_rows,
            duplicates_removed,
            missing_prices_removed,
        ) = clean_sales_data(
            sales_data
        )

        summary = create_summary(
            cleaned_data,
            original_rows,
            duplicates_removed,
            missing_prices_removed,
        )

        print_summary(summary)

        export_report(
            cleaned_data,
            summary,
            output_file,
        )

        format_excel_report(
            output_file
        )

        print(
            "\nReport created successfully."
        )

        print(
            "Report path:",
            Path(output_file).resolve(),
        )

    except Exception as error:

        print(
            "\nERROR:",
            error
        )


if __name__ == "__main__":
    main()